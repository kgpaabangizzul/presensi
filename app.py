from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, date, timedelta
import sqlite3, os, math, json, io
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm

app = Flask(__name__)
app.secret_key = 'absensi-secret-key-2024'
app.config['UPLOAD_FOLDER'] = os.path.join('static', 'uploads', 'photos')
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
DB_PATH = os.path.join('instance', 'absensi.db')

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    os.makedirs('instance', exist_ok=True)
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    conn = get_db()
    c = conn.cursor()
    c.executescript('''
        CREATE TABLE IF NOT EXISTS departemen (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nama TEXT UNIQUE NOT NULL,
            kode TEXT UNIQUE NOT NULL,
            deskripsi TEXT,
            warna TEXT DEFAULT "#3b82f6",
            aktif INTEGER DEFAULT 1,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS shift (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nama TEXT NOT NULL,
            jam_masuk TEXT NOT NULL,
            jam_keluar TEXT NOT NULL,
            toleransi_menit INTEGER DEFAULT 15,
            deskripsi TEXT,
            warna TEXT DEFAULT "#10b981",
            aktif INTEGER DEFAULT 1,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS departemen_shift (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            departemen_id INTEGER NOT NULL,
            shift_id INTEGER NOT NULL,
            FOREIGN KEY(departemen_id) REFERENCES departemen(id),
            FOREIGN KEY(shift_id) REFERENCES shift(id),
            UNIQUE(departemen_id, shift_id)
        );
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nik TEXT UNIQUE NOT NULL,
            nama TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            jabatan TEXT,
            departemen TEXT,
            departemen_id INTEGER,
            shift_id INTEGER,
            no_hp TEXT,
            alamat TEXT,
            tanggal_lahir TEXT,
            jenis_kelamin TEXT,
            foto TEXT,
            role TEXT DEFAULT "user",
            status TEXT DEFAULT "pending",
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(departemen_id) REFERENCES departemen(id),
            FOREIGN KEY(shift_id) REFERENCES shift(id)
        );
        CREATE TABLE IF NOT EXISTS absensi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            tanggal DATE NOT NULL,
            jam_masuk TIME,
            jam_keluar TIME,
            foto_masuk TEXT,
            foto_keluar TEXT,
            lat_masuk REAL,
            lng_masuk REAL,
            lat_keluar REAL,
            lng_keluar REAL,
            jarak_masuk REAL,
            jarak_keluar REAL,
            shift_id INTEGER,
            status TEXT DEFAULT "hadir",
            keterangan TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS izin (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            tanggal_mulai DATE NOT NULL,
            tanggal_selesai DATE NOT NULL,
            jenis TEXT NOT NULL,
            alasan TEXT,
            lampiran TEXT,
            status TEXT DEFAULT "pending",
            catatan_admin TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS settings (
            id INTEGER PRIMARY KEY,
            nama_perusahaan TEXT DEFAULT "PT. Absensi Digital",
            jam_masuk TEXT DEFAULT "08:00",
            jam_keluar TEXT DEFAULT "17:00",
            office_lat REAL DEFAULT -6.2088,
            office_lng REAL DEFAULT 106.8456,
            max_distance INTEGER DEFAULT 100
        );
    ''')
    c.execute("INSERT OR IGNORE INTO settings (id, nama_perusahaan) VALUES (1, 'PT. Absensi Digital')")
    admin_pass = generate_password_hash('admin123')
    c.execute("""INSERT OR IGNORE INTO users (nik,nama,email,password,jabatan,departemen,role,status)
        VALUES ('ADMIN001','Administrator','admin@absensi.com',?,'System Administrator','IT','admin','active')""", (admin_pass,))
    depts = [('IT','IT','Information Technology','#3b82f6'),('HR','HR','Human Resources','#8b5cf6'),
             ('Finance','FIN','Keuangan','#10b981'),('Marketing','MKT','Pemasaran','#f59e0b'),
             ('Operations','OPS','Operasional','#ef4444'),('Sales','SLS','Penjualan','#06b6d4')]
    for d in depts:
        c.execute("INSERT OR IGNORE INTO departemen (nama,kode,deskripsi,warna) VALUES (?,?,?,?)", d)
    shifts = [('Shift Pagi','08:00','17:00',15,'Shift normal pagi','#10b981'),
              ('Shift Siang','13:00','21:00',15,'Shift siang','#f59e0b'),
              ('Shift Malam','21:00','06:00',15,'Shift malam','#6366f1'),
              ('Shift Fleksibel','07:00','16:00',30,'Jam fleksibel','#06b6d4')]
    for s in shifts:
        c.execute("INSERT OR IGNORE INTO shift (nama,jam_masuk,jam_keluar,toleransi_menit,deskripsi,warna) VALUES (?,?,?,?,?,?)", s)
    conn.commit()
    # Auto-migration for existing databases
    migrate_cols = [
        ('users', 'departemen_id', 'INTEGER'),
        ('users', 'shift_id', 'INTEGER'),
        ('users', 'tanggal_lahir', 'TEXT'),
        ('users', 'jenis_kelamin', 'TEXT'),
        ('absensi', 'shift_id', 'INTEGER'),
        ('izin', 'catatan_admin', 'TEXT'),
    ]
    for table, col, typ in migrate_cols:
        existing = [r[1] for r in conn.execute(f'PRAGMA table_info({table})').fetchall()]
        if col not in existing:
            conn.execute(f'ALTER TABLE {table} ADD COLUMN {col} {typ}')
    conn.commit()
    conn.close()

def allowed_file(fn):
    return '.' in fn and fn.rsplit('.',1)[1].lower() in ALLOWED_EXTENSIONS

def haversine(lat1,lon1,lat2,lon2):
    R=6371000; p1,p2=math.radians(lat1),math.radians(lat2)
    dp=math.radians(lat2-lat1); dl=math.radians(lon2-lon1)
    a=math.sin(dp/2)**2+math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2*R*math.atan2(math.sqrt(a),math.sqrt(1-a))

def get_user_shift(uid, conn):
    user = conn.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if user and user['shift_id']:
        s = conn.execute("SELECT * FROM shift WHERE id=? AND aktif=1", (user['shift_id'],)).fetchone()
        if s: return dict(s)
    if user and user['departemen_id']:
        s = conn.execute("""SELECT s.* FROM shift s JOIN departemen_shift ds ON s.id=ds.shift_id
            WHERE ds.departemen_id=? AND s.aktif=1 LIMIT 1""", (user['departemen_id'],)).fetchone()
        if s: return dict(s)
    settings = conn.execute("SELECT * FROM settings WHERE id=1").fetchone()
    return {'jam_masuk': settings['jam_masuk'] if settings else '08:00',
            'jam_keluar': settings['jam_keluar'] if settings else '17:00',
            'toleransi_menit': 15, 'nama': 'Default', 'id': None}

def login_required(f):
    @wraps(f)
    def dec(*a,**kw):
        if 'user_id' not in session: return redirect(url_for('login'))
        return f(*a,**kw)
    return dec

def admin_required(f):
    @wraps(f)
    def dec(*a,**kw):
        if 'user_id' not in session or session.get('role')!='admin':
            flash('Akses ditolak!','error'); return redirect(url_for('dashboard'))
        return f(*a,**kw)
    return dec

# ---- AUTH ----
@app.route('/')
def index():
    return redirect(url_for('dashboard') if 'user_id' in session else url_for('login'))

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method=='POST':
        conn=get_db()
        user=conn.execute("SELECT * FROM users WHERE email=?",(request.form.get('email','').strip(),)).fetchone()
        conn.close()
        if user and check_password_hash(user['password'],request.form.get('password','')):
            if user['status']=='pending': flash('Akun belum divalidasi admin.','warning')
            elif user['status']=='rejected': flash('Akun ditolak. Hubungi admin.','error')
            else:
                session.update({'user_id':user['id'],'nama':user['nama'],'role':user['role'],'foto':user['foto']})
                return redirect(url_for('dashboard'))
        else: flash('Email atau password salah.','error')
    return render_template('login.html')

@app.route('/register', methods=['GET','POST'])
def register():
    conn=get_db()
    depts=conn.execute("SELECT * FROM departemen WHERE aktif=1 ORDER BY nama").fetchall()
    if request.method=='POST':
        dept_id=request.form.get('departemen_id') or None
        dept_nama=''
        if dept_id:
            d=conn.execute("SELECT nama FROM departemen WHERE id=?",(dept_id,)).fetchone()
            if d: dept_nama=d['nama']
        foto_path=None
        if 'foto' in request.files:
            f=request.files['foto']
            if f and f.filename and allowed_file(f.filename):
                nik=request.form.get('nik','new')
                fn=secure_filename(f"{nik}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{f.filename.rsplit('.',1)[1].lower()}")
                f.save(os.path.join(app.config['UPLOAD_FOLDER'],fn)); foto_path=fn
        try:
            conn.execute("""INSERT INTO users (nik,nama,email,password,jabatan,departemen,departemen_id,no_hp,alamat,tanggal_lahir,jenis_kelamin,foto)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                (request.form['nik'].strip(),request.form['nama'].strip(),request.form['email'].strip(),
                 generate_password_hash(request.form['password']),request.form.get('jabatan','').strip(),
                 dept_nama,dept_id,request.form.get('no_hp','').strip(),request.form.get('alamat','').strip(),
                 request.form.get('tanggal_lahir',''),request.form.get('jenis_kelamin',''),foto_path))
            conn.commit(); conn.close()
            flash('Registrasi berhasil! Tunggu validasi admin.','success')
            return redirect(url_for('login'))
        except: flash('NIK atau Email sudah terdaftar.','error')
    conn.close()
    return render_template('register.html',depts=depts)

@app.route('/logout')
def logout():
    session.clear(); return redirect(url_for('login'))

# ---- USER ----
@app.route('/dashboard')
@login_required
def dashboard():
    if session.get('role')=='admin': return redirect(url_for('admin_dashboard'))
    uid=session['user_id']; today=date.today().isoformat(); conn=get_db()
    absen_today=conn.execute("SELECT * FROM absensi WHERE user_id=? AND tanggal=?",(uid,today)).fetchone()
    bulan=date.today().strftime('%Y-%m')
    stats=conn.execute("""SELECT SUM(CASE WHEN status='hadir' THEN 1 ELSE 0 END) as hadir,
        SUM(CASE WHEN status='telat' THEN 1 ELSE 0 END) as telat,
        SUM(CASE WHEN status='izin' THEN 1 ELSE 0 END) as izin, COUNT(*) as total
        FROM absensi WHERE user_id=? AND strftime('%Y-%m',tanggal)=?""",(uid,bulan)).fetchone()
    riwayat=[dict(r) for r in conn.execute("""SELECT a.tanggal,a.status,a.jam_masuk,a.jam_keluar,a.jarak_masuk,s.nama as shift_nama
        FROM absensi a LEFT JOIN shift s ON a.shift_id=s.id
        WHERE a.user_id=? ORDER BY a.tanggal DESC LIMIT 30""",(uid,)).fetchall()]
    user=conn.execute("""SELECT u.*,d.nama as dept_nama,d.warna as dept_warna,
        s.nama as shift_nama,s.jam_masuk as shift_masuk,s.jam_keluar as shift_keluar
        FROM users u LEFT JOIN departemen d ON u.departemen_id=d.id
        LEFT JOIN shift s ON u.shift_id=s.id WHERE u.id=?""",(uid,)).fetchone()
    settings=conn.execute("SELECT * FROM settings WHERE id=1").fetchone()
    user_shift=get_user_shift(uid,conn); conn.close()
    return render_template('dashboard.html',absen_today=absen_today,stats=stats,
        riwayat=riwayat,user=user,settings=settings,today=today,user_shift=user_shift)

@app.route('/absen', methods=['POST'])
@login_required
def absen():
    uid=session['user_id']; today=date.today().isoformat()
    now=datetime.now().strftime('%H:%M:%S')
    lat=request.form.get('lat',type=float); lng=request.form.get('lng',type=float)
    tipe=request.form.get('tipe'); conn=get_db()
    settings=conn.execute("SELECT * FROM settings WHERE id=1").fetchone()
    off_lat=settings['office_lat'] if settings else -6.2088
    off_lng=settings['office_lng'] if settings else 106.8456
    user_shift=get_user_shift(uid,conn)
    jarak=haversine(lat,lng,off_lat,off_lng) if lat and lng else None
    max_dist=settings['max_distance'] if settings else 100
    # Blokir jika di luar radius
    if jarak is not None and jarak > max_dist:
        conn.close()
        flash(f'Absen ditolak! Anda berada {jarak:.0f}m dari kantor. Batas maksimal {max_dist}m.','error_radius')
        return redirect(url_for('dashboard'))
    foto_path=None
    if 'foto' in request.files:
        f=request.files['foto']
        if f and f.filename:
            ext=f.filename.rsplit('.',1)[-1].lower() if '.' in f.filename else 'jpg'
            fn=secure_filename(f"{uid}_{today}_{tipe}.{ext}")
            f.save(os.path.join(app.config['UPLOAD_FOLDER'],fn)); foto_path=fn
    status='hadir'; toleransi=user_shift.get('toleransi_menit',15)
    jam_shift=user_shift.get('jam_masuk','08:00')
    if tipe=='masuk':
        from datetime import time as dtime
        h,m=map(int,jam_shift.split(':'))
        batas=datetime.combine(date.today(),dtime(h,m))+timedelta(minutes=toleransi)
        if datetime.now()>batas: status='telat'
    shift_id=user_shift.get('id')
    absen_today=conn.execute("SELECT * FROM absensi WHERE user_id=? AND tanggal=?",(uid,today)).fetchone()
    if tipe=='masuk':
        if absen_today: flash('Sudah absen masuk hari ini!','warning')
        else:
            conn.execute("""INSERT INTO absensi (user_id,tanggal,jam_masuk,foto_masuk,lat_masuk,lng_masuk,jarak_masuk,shift_id,status)
                VALUES (?,?,?,?,?,?,?,?,?)""",(uid,today,now,foto_path,lat,lng,jarak,shift_id,status))
            conn.commit()
            flash(f'Absen masuk berhasil! Jarak: {jarak:.0f}m' if jarak else 'Absen masuk berhasil!','success')
    elif tipe=='keluar':
        if not absen_today: flash('Belum absen masuk!','warning')
        elif absen_today['jam_keluar']: flash('Sudah absen keluar!','warning')
        else:
            conn.execute("""UPDATE absensi SET jam_keluar=?,foto_keluar=?,lat_keluar=?,lng_keluar=?,jarak_keluar=?
                WHERE user_id=? AND tanggal=?""",(now,foto_path,lat,lng,jarak,uid,today))
            conn.commit(); flash('Absen keluar berhasil!','success')
    conn.close(); return redirect(url_for('dashboard'))

@app.route('/riwayat')
@login_required
def riwayat():
    uid=session['user_id']; bulan=request.args.get('bulan',date.today().strftime('%Y-%m'))
    conn=get_db()
    data=conn.execute("""SELECT a.*,s.nama as shift_nama,s.jam_masuk as shift_masuk
        FROM absensi a LEFT JOIN shift s ON a.shift_id=s.id
        WHERE a.user_id=? AND strftime('%Y-%m',a.tanggal)=? ORDER BY a.tanggal DESC""",(uid,bulan)).fetchall()
    conn.close(); return render_template('riwayat.html',data=data,bulan=bulan)

@app.route('/izin', methods=['GET','POST'])
@login_required
def izin():
    uid=session['user_id']
    if request.method=='POST':
        conn=get_db(); lamp=None
        if 'lampiran' in request.files:
            f=request.files['lampiran']
            if f and f.filename and allowed_file(f.filename):
                fn=secure_filename(f"{uid}_izin_{datetime.now().strftime('%Y%m%d%H%M%S')}.{f.filename.rsplit('.',1)[-1]}")
                f.save(os.path.join(app.config['UPLOAD_FOLDER'],fn)); lamp=fn
        conn.execute("INSERT INTO izin (user_id,tanggal_mulai,tanggal_selesai,jenis,alasan,lampiran) VALUES (?,?,?,?,?,?)",
            (uid,request.form['tanggal_mulai'],request.form['tanggal_selesai'],request.form['jenis'],request.form['alasan'],lamp))
        conn.commit(); conn.close(); flash('Permohonan izin berhasil dikirim!','success')
        return redirect(url_for('izin'))
    conn=get_db()
    data=conn.execute("SELECT * FROM izin WHERE user_id=? ORDER BY created_at DESC",(uid,)).fetchall()
    conn.close(); return render_template('izin.html',data=data)

@app.route('/profil', methods=['GET','POST'])
@login_required
def profil():
    uid=session['user_id']; conn=get_db()
    if request.method=='POST':
        no_hp=request.form.get('no_hp',''); alamat=request.form.get('alamat',''); foto_path=None
        if 'foto' in request.files:
            f=request.files['foto']
            if f and f.filename and allowed_file(f.filename):
                ext=f.filename.rsplit('.',1)[-1].lower()
                fn=secure_filename(f"user_{uid}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}")
                f.save(os.path.join(app.config['UPLOAD_FOLDER'],fn)); foto_path=fn
        if foto_path:
            conn.execute("UPDATE users SET no_hp=?,alamat=?,foto=? WHERE id=?",(no_hp,alamat,foto_path,uid))
            session['foto']=foto_path
        else: conn.execute("UPDATE users SET no_hp=?,alamat=? WHERE id=?",(no_hp,alamat,uid))
        conn.commit(); flash('Profil berhasil diperbarui!','success')
    user=conn.execute("""SELECT u.*,d.nama as dept_nama,d.warna as dept_warna,
        s.nama as shift_nama,s.jam_masuk as shift_masuk,s.jam_keluar as shift_keluar
        FROM users u LEFT JOIN departemen d ON u.departemen_id=d.id
        LEFT JOIN shift s ON u.shift_id=s.id WHERE u.id=?""",(uid,)).fetchone()
    conn.close(); return render_template('profil.html',user=user)

# ---- ADMIN DASHBOARD ----
@app.route('/admin')
@admin_required
def admin_dashboard():
    conn=get_db(); today=date.today().isoformat()
    total_user=conn.execute("SELECT COUNT(*) FROM users WHERE role='user' AND status='active'").fetchone()[0]
    pending=conn.execute("SELECT COUNT(*) FROM users WHERE status='pending'").fetchone()[0]
    hadir_today=conn.execute("SELECT COUNT(*) FROM absensi WHERE tanggal=? AND status IN ('hadir','telat')",(today,)).fetchone()[0]
    telat_today=conn.execute("SELECT COUNT(*) FROM absensi WHERE tanggal=? AND status='telat'",(today,)).fetchone()[0]
    izin_pending=conn.execute("SELECT COUNT(*) FROM izin WHERE status='pending'").fetchone()[0]
    total_dept=conn.execute("SELECT COUNT(*) FROM departemen WHERE aktif=1").fetchone()[0]
    total_shift=conn.execute("SELECT COUNT(*) FROM shift WHERE aktif=1").fetchone()[0]
    chart_data=[{'tanggal':(date.today()-timedelta(days=i)).isoformat(),
        'hadir':conn.execute("SELECT COUNT(*) FROM absensi WHERE tanggal=? AND status IN ('hadir','telat')",
            ((date.today()-timedelta(days=i)).isoformat(),)).fetchone()[0],
        'telat':conn.execute("SELECT COUNT(*) FROM absensi WHERE tanggal=? AND status='telat'",
            ((date.today()-timedelta(days=i)).isoformat(),)).fetchone()[0]}
        for i in range(6,-1,-1)]
    dept_stats=[dict(r) for r in conn.execute("""
        SELECT d.nama as departemen, d.warna, d.kode,
            COUNT(DISTINCT u.id) as total_pegawai,
            SUM(CASE WHEN a.tanggal=? AND a.status IN ('hadir','telat') THEN 1 ELSE 0 END) as hadir_today
        FROM departemen d LEFT JOIN users u ON u.departemen_id=d.id AND u.status='active' AND u.role='user'
        LEFT JOIN absensi a ON u.id=a.user_id AND a.tanggal=?
        WHERE d.aktif=1 GROUP BY d.id ORDER BY total_pegawai DESC""",(today,today)).fetchall()]
    recent=conn.execute("""SELECT a.*,u.nama,u.jabatan,u.foto,u.departemen,s.nama as shift_nama
        FROM absensi a JOIN users u ON a.user_id=u.id LEFT JOIN shift s ON a.shift_id=s.id
        WHERE a.tanggal=? ORDER BY a.jam_masuk DESC LIMIT 10""",(today,)).fetchall()
    recent_depts=conn.execute("SELECT *,(SELECT COUNT(*) FROM users WHERE departemen_id=departemen.id AND status='active') as jml FROM departemen WHERE aktif=1 ORDER BY nama").fetchall()
    recent_shifts=conn.execute("SELECT *,(SELECT COUNT(*) FROM users WHERE shift_id=shift.id AND status='active') as jml FROM shift WHERE aktif=1 ORDER BY jam_masuk").fetchall()
    conn.close()
    return render_template('admin/dashboard.html',total_user=total_user,pending=pending,
        hadir_today=hadir_today,telat_today=telat_today,izin_pending=izin_pending,
        total_dept=total_dept,total_shift=total_shift,chart_data=json.dumps(chart_data),
        dept_stats=dept_stats,recent=recent,recent_depts=recent_depts,recent_shifts=recent_shifts,today=today)

# ---- ADMIN DEPARTEMEN ----
@app.route('/admin/departemen')
@admin_required
def admin_departemen():
    conn=get_db()
    depts=conn.execute("""SELECT d.*,COUNT(DISTINCT u.id) as total_pegawai,COUNT(DISTINCT ds.shift_id) as total_shift
        FROM departemen d LEFT JOIN users u ON u.departemen_id=d.id AND u.status='active' AND u.role='user'
        LEFT JOIN departemen_shift ds ON ds.departemen_id=d.id
        GROUP BY d.id ORDER BY d.nama""").fetchall()
    shifts=conn.execute("SELECT * FROM shift WHERE aktif=1 ORDER BY jam_masuk").fetchall()
    dept_shifts={} 
    for ds in conn.execute("SELECT * FROM departemen_shift").fetchall():
        dept_shifts.setdefault(ds['departemen_id'],[]).append(ds['shift_id'])
    conn.close()
    return render_template('admin/departemen.html',depts=depts,shifts=shifts,dept_shifts=dept_shifts)

@app.route('/admin/departemen/tambah', methods=['POST'])
@admin_required
def tambah_departemen():
    conn=get_db()
    try:
        conn.execute("INSERT INTO departemen (nama,kode,deskripsi,warna) VALUES (?,?,?,?)",
            (request.form['nama'],request.form['kode'].upper(),request.form.get('deskripsi',''),request.form.get('warna','#3b82f6')))
        conn.commit(); flash(f'Departemen "{request.form["nama"]}" ditambahkan!','success')
    except: flash('Nama atau kode sudah ada!','error')
    conn.close(); return redirect(url_for('admin_departemen'))

@app.route('/admin/departemen/edit/<int:did>', methods=['POST'])
@admin_required
def edit_departemen(did):
    conn=get_db()
    try:
        conn.execute("UPDATE departemen SET nama=?,kode=?,deskripsi=?,warna=?,aktif=? WHERE id=?",
            (request.form['nama'],request.form['kode'].upper(),request.form.get('deskripsi',''),
             request.form.get('warna','#3b82f6'),1 if request.form.get('aktif') else 0,did))
        conn.commit(); flash('Departemen diperbarui!','success')
    except Exception as e: flash('Gagal: '+str(e),'error')
    conn.close(); return redirect(url_for('admin_departemen'))

@app.route('/admin/departemen/hapus/<int:did>', methods=['POST'])
@admin_required
def hapus_departemen(did):
    conn=get_db()
    peg=conn.execute("SELECT COUNT(*) FROM users WHERE departemen_id=?",(did,)).fetchone()[0]
    if peg>0: flash(f'Tidak bisa hapus: masih ada {peg} pegawai!','error')
    else:
        conn.execute("DELETE FROM departemen_shift WHERE departemen_id=?",(did,))
        conn.execute("DELETE FROM departemen WHERE id=?",(did,))
        conn.commit(); flash('Departemen dihapus!','success')
    conn.close(); return redirect(url_for('admin_departemen'))

@app.route('/admin/departemen/<int:did>/shift', methods=['POST'])
@admin_required
def atur_shift_departemen(did):
    conn=get_db()
    conn.execute("DELETE FROM departemen_shift WHERE departemen_id=?",(did,))
    for sid in request.form.getlist('shift_ids'):
        conn.execute("INSERT OR IGNORE INTO departemen_shift (departemen_id,shift_id) VALUES (?,?)",(did,sid))
    conn.commit(); conn.close()
    flash('Shift departemen diperbarui!','success'); return redirect(url_for('admin_departemen'))

# ---- ADMIN SHIFT ----
@app.route('/admin/shift')
@admin_required
def admin_shift():
    conn=get_db()
    shifts=conn.execute("""SELECT s.*,COUNT(DISTINCT u.id) as total_pegawai,COUNT(DISTINCT ds.departemen_id) as total_dept
        FROM shift s LEFT JOIN users u ON u.shift_id=s.id AND u.status='active'
        LEFT JOIN departemen_shift ds ON ds.shift_id=s.id
        GROUP BY s.id ORDER BY s.jam_masuk""").fetchall()
    depts=conn.execute("SELECT * FROM departemen WHERE aktif=1 ORDER BY nama").fetchall()
    conn.close(); return render_template('admin/shift.html',shifts=shifts,depts=depts)

@app.route('/admin/shift/tambah', methods=['POST'])
@admin_required
def tambah_shift():
    conn=get_db()
    try:
        conn.execute("INSERT INTO shift (nama,jam_masuk,jam_keluar,toleransi_menit,deskripsi,warna) VALUES (?,?,?,?,?,?)",
            (request.form['nama'],request.form['jam_masuk'],request.form['jam_keluar'],
             int(request.form.get('toleransi_menit',15)),request.form.get('deskripsi',''),request.form.get('warna','#10b981')))
        conn.commit(); flash(f'Shift "{request.form["nama"]}" ditambahkan!','success')
    except Exception as e: flash('Gagal: '+str(e),'error')
    conn.close(); return redirect(url_for('admin_shift'))

@app.route('/admin/shift/edit/<int:sid>', methods=['POST'])
@admin_required
def edit_shift(sid):
    conn=get_db()
    conn.execute("UPDATE shift SET nama=?,jam_masuk=?,jam_keluar=?,toleransi_menit=?,deskripsi=?,warna=?,aktif=? WHERE id=?",
        (request.form['nama'],request.form['jam_masuk'],request.form['jam_keluar'],
         int(request.form.get('toleransi_menit',15)),request.form.get('deskripsi',''),
         request.form.get('warna','#10b981'),1 if request.form.get('aktif') else 0,sid))
    conn.commit(); conn.close(); flash('Shift diperbarui!','success')
    return redirect(url_for('admin_shift'))

@app.route('/admin/shift/hapus/<int:sid>', methods=['POST'])
@admin_required
def hapus_shift(sid):
    conn=get_db()
    peg=conn.execute("SELECT COUNT(*) FROM users WHERE shift_id=?",(sid,)).fetchone()[0]
    if peg>0: flash(f'Tidak bisa hapus: {peg} pegawai masih menggunakan shift ini!','error')
    else:
        conn.execute("DELETE FROM departemen_shift WHERE shift_id=?",(sid,))
        conn.execute("DELETE FROM shift WHERE id=?",(sid,))
        conn.commit(); flash('Shift dihapus!','success')
    conn.close(); return redirect(url_for('admin_shift'))

# ---- ADMIN PEGAWAI ----
@app.route('/admin/pegawai')
@admin_required
def admin_pegawai():
    conn=get_db(); q=request.args.get('q',''); sf=request.args.get('status',''); df=request.args.get('dept','')
    query="""SELECT u.*,d.nama as dept_nama,d.warna as dept_warna,s.nama as shift_nama,
        s.jam_masuk as shift_masuk,s.jam_keluar as shift_keluar FROM users u
        LEFT JOIN departemen d ON u.departemen_id=d.id LEFT JOIN shift s ON u.shift_id=s.id WHERE u.role='user'"""
    params=[]
    if q:
        query+=" AND (u.nama LIKE ? OR u.nik LIKE ? OR u.email LIKE ?)"
        params.extend([f'%{q}%']*3)
    if sf: query+=" AND u.status=?"; params.append(sf)
    if df: query+=" AND u.departemen_id=?"; params.append(df)
    query+=" ORDER BY u.created_at DESC"
    users=conn.execute(query,params).fetchall()
    depts=conn.execute("SELECT * FROM departemen WHERE aktif=1 ORDER BY nama").fetchall()
    shifts=conn.execute("SELECT * FROM shift WHERE aktif=1 ORDER BY jam_masuk").fetchall()
    stats={'total':len(users),'active':sum(1 for u in users if u['status']=='active'),
           'pending':sum(1 for u in users if u['status']=='pending')}
    conn.close()
    return render_template('admin/pegawai.html',users=users,q=q,status_filter=sf,dept_filter=df,
        depts=depts,shifts=shifts,stats=stats)

@app.route('/admin/pegawai/tambah', methods=['POST'])
@admin_required
def tambah_pegawai():
    conn=get_db()
    dept_id=request.form.get('departemen_id') or None; dept_nama=''
    if dept_id:
        d=conn.execute("SELECT nama FROM departemen WHERE id=?",(dept_id,)).fetchone()
        if d: dept_nama=d['nama']
    foto_path=None
    if 'foto' in request.files:
        f=request.files['foto']
        if f and f.filename and allowed_file(f.filename):
            fn=secure_filename(f"{request.form.get('nik','new')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{f.filename.rsplit('.',1)[-1].lower()}")
            f.save(os.path.join(app.config['UPLOAD_FOLDER'],fn)); foto_path=fn
    try:
        conn.execute("""INSERT INTO users (nik,nama,email,password,jabatan,departemen,departemen_id,shift_id,no_hp,alamat,tanggal_lahir,jenis_kelamin,foto,status)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (request.form['nik'],request.form['nama'],request.form['email'],
             generate_password_hash(request.form['password']),request.form.get('jabatan',''),
             dept_nama,dept_id,request.form.get('shift_id') or None,
             request.form.get('no_hp',''),request.form.get('alamat',''),
             request.form.get('tanggal_lahir',''),request.form.get('jenis_kelamin',''),
             foto_path,request.form.get('status','active')))
        conn.commit(); flash('Pegawai berhasil ditambahkan!','success')
    except Exception as e: flash('NIK atau Email sudah ada: '+str(e),'error')
    conn.close(); return redirect(url_for('admin_pegawai'))

@app.route('/admin/pegawai/edit/<int:uid>', methods=['GET','POST'])
@admin_required
def edit_pegawai(uid):
    conn=get_db()
    if request.method=='POST':
        dept_id=request.form.get('departemen_id') or None; dept_nama=''
        if dept_id:
            d=conn.execute("SELECT nama FROM departemen WHERE id=?",(dept_id,)).fetchone()
            if d: dept_nama=d['nama']
        foto_path=None
        if 'foto' in request.files:
            f=request.files['foto']
            if f and f.filename and allowed_file(f.filename):
                fn=secure_filename(f"user_{uid}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{f.filename.rsplit('.',1)[-1].lower()}")
                f.save(os.path.join(app.config['UPLOAD_FOLDER'],fn)); foto_path=fn
        fields=["nik=?","nama=?","email=?","jabatan=?","departemen=?","departemen_id=?","shift_id=?",
                "no_hp=?","alamat=?","tanggal_lahir=?","jenis_kelamin=?","status=?"]
        params=[request.form['nik'],request.form['nama'],request.form['email'],
                request.form.get('jabatan',''),dept_nama,dept_id,request.form.get('shift_id') or None,
                request.form.get('no_hp',''),request.form.get('alamat',''),
                request.form.get('tanggal_lahir',''),request.form.get('jenis_kelamin',''),
                request.form.get('status','active')]
        if foto_path: fields.append("foto=?"); params.append(foto_path)
        if request.form.get('password'): fields.append("password=?"); params.append(generate_password_hash(request.form['password']))
        params.append(uid)
        conn.execute(f"UPDATE users SET {','.join(fields)} WHERE id=?",params)
        conn.commit(); conn.close(); flash('Data pegawai diperbarui!','success')
        return redirect(url_for('admin_pegawai'))
    user=conn.execute("""SELECT u.*,d.nama as dept_nama,s.nama as shift_nama
        FROM users u LEFT JOIN departemen d ON u.departemen_id=d.id
        LEFT JOIN shift s ON u.shift_id=s.id WHERE u.id=?""",(uid,)).fetchone()
    depts=conn.execute("SELECT * FROM departemen WHERE aktif=1 ORDER BY nama").fetchall()
    shifts=conn.execute("SELECT * FROM shift WHERE aktif=1 ORDER BY jam_masuk").fetchall()
    conn.close(); return render_template('admin/edit_pegawai.html',user=user,depts=depts,shifts=shifts)

@app.route('/admin/pegawai/hapus/<int:uid>', methods=['POST'])
@admin_required
def hapus_pegawai(uid):
    conn=get_db()
    cnt=conn.execute("SELECT COUNT(*) FROM absensi WHERE user_id=?",(uid,)).fetchone()[0]
    if cnt>0:
        conn.execute("UPDATE users SET status='rejected' WHERE id=?",(uid,))
        conn.commit(); flash('Pegawai dinonaktifkan (ada data absensi).','warning')
    else:
        conn.execute("DELETE FROM izin WHERE user_id=?",(uid,))
        conn.execute("DELETE FROM users WHERE id=?",(uid,))
        conn.commit(); flash('Pegawai dihapus!','success')
    conn.close(); return redirect(url_for('admin_pegawai'))

@app.route('/admin/validasi/<int:uid>/<action>')
@admin_required
def validasi_user(uid,action):
    conn=get_db()
    conn.execute("UPDATE users SET status=? WHERE id=?",('active' if action=='approve' else 'rejected',uid))
    conn.commit(); conn.close()
    flash('Akun disetujui!' if action=='approve' else 'Akun ditolak!','success' if action=='approve' else 'info')
    return redirect(url_for('admin_pegawai'))

# ---- ADMIN ABSENSI ----
@app.route('/admin/absensi')
@admin_required
def admin_absensi():
    conn=get_db(); bulan=request.args.get('bulan',date.today().strftime('%Y-%m')); dept=request.args.get('dept','')
    q=f"SELECT a.*,u.nama,u.nik,u.jabatan,u.departemen,u.foto,s.nama as shift_nama FROM absensi a JOIN users u ON a.user_id=u.id LEFT JOIN shift s ON a.shift_id=s.id WHERE strftime('%Y-%m',a.tanggal)=?"
    params=[bulan]
    if dept: q+=" AND u.departemen_id=?"; params.append(dept)
    q+=" ORDER BY a.tanggal DESC,a.jam_masuk DESC"
    data=conn.execute(q,params).fetchall()
    depts=conn.execute("SELECT * FROM departemen WHERE aktif=1 ORDER BY nama").fetchall()
    conn.close(); return render_template('admin/absensi.html',data=data,bulan=bulan,dept=dept,depts=depts)

# ---- ADMIN IZIN ----
@app.route('/admin/izin')
@admin_required
def admin_izin():
    conn=get_db()
    data=conn.execute("SELECT i.*,u.nama,u.nik,u.departemen FROM izin i JOIN users u ON i.user_id=u.id ORDER BY i.created_at DESC").fetchall()
    conn.close(); return render_template('admin/izin.html',data=data)

@app.route('/admin/izin/<int:iid>/<action>')
@admin_required
def proses_izin(iid,action):
    conn=get_db(); iz=conn.execute("SELECT * FROM izin WHERE id=?",(iid,)).fetchone()
    if iz and action in ['approve','reject']:
        conn.execute("UPDATE izin SET status=? WHERE id=?",('approved' if action=='approve' else 'rejected',iid))
        if action=='approve':
            d1=datetime.strptime(iz['tanggal_mulai'],'%Y-%m-%d'); d2=datetime.strptime(iz['tanggal_selesai'],'%Y-%m-%d'); cur=d1
            while cur<=d2:
                conn.execute("INSERT OR IGNORE INTO absensi (user_id,tanggal,status,keterangan) VALUES (?,?,?,?)",
                    (iz['user_id'],cur.date().isoformat(),'izin',iz['jenis'])); cur+=timedelta(days=1)
        conn.commit(); flash('Izin diproses!','success')
    conn.close(); return redirect(url_for('admin_izin'))

# ---- ADMIN LAPORAN ----
@app.route('/admin/laporan')
@admin_required
def admin_laporan():
    conn=get_db(); bulan=request.args.get('bulan',date.today().strftime('%Y-%m'))
    users=conn.execute("""SELECT u.*,d.nama as dept_nama,s.nama as shift_nama FROM users u
        LEFT JOIN departemen d ON u.departemen_id=d.id LEFT JOIN shift s ON u.shift_id=s.id
        WHERE u.role='user' AND u.status='active' ORDER BY u.nama""").fetchall()
    rekap=[{'user':u,'stats':conn.execute("""SELECT SUM(CASE WHEN status='hadir' THEN 1 ELSE 0 END) as hadir,
        SUM(CASE WHEN status='telat' THEN 1 ELSE 0 END) as telat,SUM(CASE WHEN status='izin' THEN 1 ELSE 0 END) as izin,
        COUNT(*) as total FROM absensi WHERE user_id=? AND strftime('%Y-%m',tanggal)=?""",(u['id'],bulan)).fetchone()} for u in users]
    conn.close(); return render_template('admin/laporan.html',rekap=rekap,bulan=bulan)

@app.route('/admin/export/excel')
@admin_required
def export_excel():
    bulan=request.args.get('bulan',date.today().strftime('%Y-%m')); conn=get_db()
    data=conn.execute("""SELECT a.tanggal,u.nik,u.nama,u.departemen,u.jabatan,a.jam_masuk,a.jam_keluar,
        a.jarak_masuk,a.status,a.keterangan,s.nama as shift_nama
        FROM absensi a JOIN users u ON a.user_id=u.id LEFT JOIN shift s ON a.shift_id=s.id
        WHERE strftime('%Y-%m',a.tanggal)=? ORDER BY a.tanggal,u.nama""",(bulan,)).fetchall()
    settings=conn.execute("SELECT * FROM settings WHERE id=1").fetchone(); conn.close()
    wb=openpyxl.Workbook(); ws=wb.active; ws.title=f"Absensi {bulan}"
    for col,w in zip('ABCDEFGHIJK',[12,14,22,18,20,12,12,14,10,15,20]): ws.column_dimensions[col].width=w
    ws.merge_cells('A1:K1'); ws['A1']=f"LAPORAN ABSENSI - {settings['nama_perusahaan'] if settings else 'PT Absensi'}"
    ws['A1'].font=Font(bold=True,size=14); ws['A1'].alignment=Alignment(horizontal='center')
    ws.merge_cells('A2:K2'); ws['A2']=f"Periode: {bulan}"; ws['A2'].alignment=Alignment(horizontal='center')
    for col,h in enumerate(['Tanggal','NIK','Nama','Departemen','Jabatan','Jam Masuk','Jam Keluar','Jarak (m)','Status','Shift','Keterangan'],1):
        cell=ws.cell(4,col,h); cell.font=Font(bold=True,color='FFFFFF')
        cell.fill=PatternFill(fill_type='solid',fgColor='1E3A5F'); cell.alignment=Alignment(horizontal='center')
    sc={'hadir':'C8E6C9','telat':'FFE082','izin':'BBDEFB','alpha':'FFCDD2'}
    for ri,row in enumerate(data,5):
        for col,val in enumerate([str(row['tanggal']),row['nik'],row['nama'],row['departemen'] or '-',
            row['jabatan'] or '-',str(row['jam_masuk'] or '-'),str(row['jam_keluar'] or '-'),
            f"{row['jarak_masuk']:.0f}" if row['jarak_masuk'] else '-',row['status'],
            row['shift_nama'] or '-',row['keterangan'] or ''],1):
            cell=ws.cell(ri,col,val); cell.fill=PatternFill(fill_type='solid',fgColor=sc.get(row['status'],'FFFFFF'))
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,as_attachment=True,download_name=f"absensi_{bulan}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/admin/export/pdf')
@admin_required
def export_pdf():
    bulan=request.args.get('bulan',date.today().strftime('%Y-%m')); conn=get_db()
    data=conn.execute("""SELECT a.tanggal,u.nik,u.nama,u.departemen,a.jam_masuk,a.jam_keluar,
        a.jarak_masuk,a.status,s.nama as shift_nama FROM absensi a JOIN users u ON a.user_id=u.id
        LEFT JOIN shift s ON a.shift_id=s.id WHERE strftime('%Y-%m',a.tanggal)=? ORDER BY a.tanggal,u.nama""",(bulan,)).fetchall()
    settings=conn.execute("SELECT * FROM settings WHERE id=1").fetchone(); conn.close()
    out=io.BytesIO()
    doc=SimpleDocTemplate(out,pagesize=landscape(A4),rightMargin=1*cm,leftMargin=1*cm,topMargin=2*cm,bottomMargin=1*cm)
    el=[Paragraph(f"LAPORAN ABSENSI - {settings['nama_perusahaan'] if settings else 'PT Absensi'}",
            ParagraphStyle('T',fontSize=14,spaceAfter=4,fontName='Helvetica-Bold',alignment=1)),
        Paragraph(f"Periode: {bulan}",ParagraphStyle('S',fontSize=10,spaceAfter=10,alignment=1)),Spacer(1,0.3*cm)]
    td=[['No','Tanggal','NIK','Nama','Dept','Shift','Masuk','Keluar','Jarak','Status']]
    for i,row in enumerate(data,1):
        td.append([i,str(row['tanggal']),row['nik'],row['nama'],row['departemen'] or '-',
            row['shift_nama'] or '-',str(row['jam_masuk'] or '-'),str(row['jam_keluar'] or '-'),
            f"{row['jarak_masuk']:.0f}m" if row['jarak_masuk'] else '-',row['status']])
    t=Table(td,colWidths=[0.8*cm,2.2*cm,2.2*cm,3.5*cm,2.5*cm,2.5*cm,2*cm,2*cm,1.8*cm,1.8*cm])
    sc2={'hadir':colors.HexColor('#C8E6C9'),'telat':colors.HexColor('#FFE082'),'izin':colors.HexColor('#BBDEFB')}
    sty=[('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1E3A5F')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
         ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),7),
         ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F5F5F5')]),
         ('GRID',(0,0),(-1,-1),0.5,colors.grey),('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE')]
    for i,row in enumerate(data,1):
        c=sc2.get(row['status'])
        if c: sty.append(('BACKGROUND',(9,i),(9,i),c))
    t.setStyle(TableStyle(sty)); el.append(t); doc.build(el); out.seek(0)
    return send_file(out,as_attachment=True,download_name=f"absensi_{bulan}.pdf",mimetype='application/pdf')

# ---- ADMIN GRAFIK ----
@app.route('/admin/grafik')
@admin_required
def admin_grafik():
    conn=get_db(); bulan=request.args.get('bulan',date.today().strftime('%Y-%m'))
    harian=conn.execute("""SELECT tanggal,SUM(CASE WHEN status='hadir' THEN 1 ELSE 0 END) as hadir,
        SUM(CASE WHEN status='telat' THEN 1 ELSE 0 END) as telat,SUM(CASE WHEN status='izin' THEN 1 ELSE 0 END) as izin
        FROM absensi WHERE strftime('%Y-%m',tanggal)=? GROUP BY tanggal ORDER BY tanggal""",(bulan,)).fetchall()
    dept=conn.execute("""SELECT d.nama as departemen,d.warna,
        SUM(CASE WHEN a.status='hadir' THEN 1 ELSE 0 END) as hadir,
        SUM(CASE WHEN a.status='telat' THEN 1 ELSE 0 END) as telat,
        SUM(CASE WHEN a.status='izin' THEN 1 ELSE 0 END) as izin,COUNT(a.id) as total
        FROM departemen d LEFT JOIN users u ON u.departemen_id=d.id AND u.role='user' AND u.status='active'
        LEFT JOIN absensi a ON u.id=a.user_id AND strftime('%Y-%m',a.tanggal)=?
        WHERE d.aktif=1 GROUP BY d.id""",(bulan,)).fetchall()
    conn.close()
    return render_template('admin/grafik.html',harian=json.dumps([dict(r) for r in harian]),
        dept=json.dumps([dict(r) for r in dept]),bulan=bulan)

# ---- ADMIN SETTINGS ----
@app.route('/admin/settings', methods=['GET','POST'])
@admin_required
def admin_settings():
    conn=get_db()
    if request.method=='POST':
        conn.execute("UPDATE settings SET nama_perusahaan=?,jam_masuk=?,jam_keluar=?,office_lat=?,office_lng=?,max_distance=? WHERE id=1",
            (request.form['nama_perusahaan'],request.form['jam_masuk'],request.form['jam_keluar'],
             float(request.form['office_lat']),float(request.form['office_lng']),int(request.form['max_distance'])))
        conn.commit(); flash('Settings disimpan!','success')
    settings=conn.execute("SELECT * FROM settings WHERE id=1").fetchone()
    conn.close(); return render_template('admin/settings.html',settings=settings)

@app.route('/api/shift-by-dept/<int:dept_id>')
@login_required
def api_shift_by_dept(dept_id):
    conn=get_db()
    shifts=conn.execute("""SELECT s.* FROM shift s JOIN departemen_shift ds ON s.id=ds.shift_id
        WHERE ds.departemen_id=? AND s.aktif=1""",(dept_id,)).fetchall()
    conn.close(); return jsonify([dict(s) for s in shifts])

if __name__=='__main__':
    init_db(); app.run(debug=True, host='0.0.0.0',port=5000)
